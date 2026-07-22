"""
Загружает каталог позиций, тех.карты и статьи обучения из seed_products.xlsx в базу.
Запускать заново каждый раз, когда меняете точки/категории/позиции в Excel.
(Если каталог уже отредактирован через бота — повторный запуск ПЕРЕЗАПИШЕТ
его содержимым Excel-файла для точек/позиций; рецепты и статьи обучения —
добавляются, уже существующие с тем же названием пропускаются, см. ниже.)

Использование:
    python import_products.py [путь_к_файлу.xlsx]
"""
import re
import sys
import openpyxl
import db


def parse_ingredient_line(line: str):
    """«Свёкла — 200 г» -> (Свёкла, 200.0, г). «Соль — по вкусу» -> (Соль, None, по вкусу)."""
    line = line.strip()
    if not line:
        return None
    parts = re.split(r"\s+—\s+|\s+-\s+", line, maxsplit=1)
    if len(parts) != 2:
        return (line, None, None)
    name, rest = parts[0].strip(), parts[1].strip()
    m = re.match(r"^([\d.,]+)\s*(.*)$", rest)
    if m:
        amount = float(m.group(1).replace(",", "."))
        unit = m.group(2).strip() or None
        return (name, amount, unit)
    return (name, None, rest or None)


def format_method(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return ""
    if all(re.match(r"^\d+[.)]", l) for l in lines):
        return "\n".join(lines)
    return "\n".join(f"{i}. {l}" for i, l in enumerate(lines, 1))


def import_recipes(wb) -> dict:
    if "Рецепты" not in wb.sheetnames:
        return {"added": 0, "skipped": 0}
    ws = wb["Рецепты"]
    existing = {(r["name"], r["category"] or "") for r in db.list_recipes()}
    added = skipped = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        name, category, ingredients_text, method_text = (row + (None,) * 4)[:4]
        if not name:
            continue
        name = str(name).strip()
        category = str(category).strip() if category else ""
        if (name, category) in existing:
            print(f"  ! Рецепт «{name}» ({category or 'без категории'}) уже есть — пропущен")
            skipped += 1
            continue
        recipe_id = db.create_recipe(name, category, format_method(str(method_text or "")))
        if ingredients_text:
            for i, line in enumerate(str(ingredients_text).splitlines()):
                parsed = parse_ingredient_line(line)
                if parsed:
                    ing_name, amount, unit = parsed
                    db.add_recipe_ingredient(recipe_id, ing_name, amount, unit, i)
        print(f"  Рецепт «{name}» добавлен")
        added += 1
    return {"added": added, "skipped": skipped}


def import_kb(wb) -> dict:
    if "Обучение" not in wb.sheetnames:
        return {"added": 0, "skipped": 0}
    ws = wb["Обучение"]
    added = skipped = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        section, title, body = (row + (None,) * 3)[:3]
        if not section or not title:
            continue
        section, title = str(section).strip(), str(title).strip()
        existing = [a for a in db.list_kb_articles(section) if a["title"] == title]
        if existing:
            print(f"  ! Статья «{title}» ({section}) уже есть — пропущена")
            skipped += 1
            continue
        db.create_kb_article(section, title, str(body or "").strip())
        print(f"  Статья «{title}» добавлена в «{section}»")
        added += 1
    return {"added": added, "skipped": skipped}


def main(path: str = "seed_products.xlsx"):
    db.init_db()
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Лист "Точки" ---
    ws_points = wb["Точки"]
    point_ids = {}
    for row in ws_points.iter_rows(min_row=5, values_only=True):
        name, chat_id, header_template = (row + (None, None, None))[:3]
        if not name:
            continue
        chat_id = int(chat_id) if chat_id not in (None, "") else None
        header_template = header_template or "📦 Заказ — {point}, {date}"
        point_id = db.upsert_point(str(name).strip(), chat_id, str(header_template).strip())
        point_ids[str(name).strip()] = point_id
        db.clear_products_for_point(point_id)   # products first (FK -> categories)
        db.clear_categories_for_point(point_id)
        print(f"Точка «{name}» -> id={point_id} (chat_id={chat_id})")

    # --- Лист "Позиции" ---
    ws_prod = wb["Позиции"]
    sort_counters = {}
    added = 0
    skipped = []
    for row in ws_prod.iter_rows(min_row=5, values_only=True):
        point_name, category, name, unit, min_qty, max_qty = (row + (None,) * 6)[:6]
        if not point_name or not name:
            continue
        point_name = str(point_name).strip()
        if point_name not in point_ids:
            print(f"  ! Точка «{point_name}» не найдена на листе «Точки» — строка пропущена: {name}")
            skipped.append(f"«{name}» — точка «{point_name}» не найдена")
            continue
        if min_qty is None or max_qty is None:
            print(f"  ! Пропущена позиция без мин/макс: {name}")
            skipped.append(f"«{name}» — не указан мин/макс")
            continue
        point_id = point_ids[point_name]
        category_name = str(category).strip() if category else "БЕЗ КАТЕГОРИИ"
        category_id = db.get_or_create_category(point_id, category_name)

        key = (point_id, category_id)
        sort_counters.setdefault(key, 0)
        db.add_product(
            point_id=point_id,
            category_id=category_id,
            name=str(name).strip(),
            unit=str(unit).strip() if unit else "шт",
            min_qty=float(min_qty),
            max_qty=float(max_qty),
            sort_order=sort_counters[key],
        )
        sort_counters[key] += 1
        added += 1

    print(f"\nГотово. Загружено позиций: {added}")

    # --- Листы "Рецепты" и "Обучение" (опциональные) ---
    recipes_result = import_recipes(wb)
    if recipes_result["added"] or recipes_result["skipped"]:
        print(f"Рецептов добавлено: {recipes_result['added']}, пропущено (уже есть): {recipes_result['skipped']}")

    kb_result = import_kb(wb)
    if kb_result["added"] or kb_result["skipped"]:
        print(f"Статей обучения добавлено: {kb_result['added']}, пропущено (уже есть): {kb_result['skipped']}")

    return {
        "points": len(point_ids),
        "products": added,
        "skipped": skipped,
        "recipes_added": recipes_result["added"],
        "recipes_skipped": recipes_result["skipped"],
        "kb_added": kb_result["added"],
        "kb_skipped": kb_result["skipped"],
    }


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "seed_products.xlsx"
    main(path)
